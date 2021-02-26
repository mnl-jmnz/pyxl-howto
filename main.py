#!/usr/bin/env python

import sys
from openpyxl import Workbook

def main(args):
    output_excel_file = "sample.xlsx"

    if len(args) > 1:
        output_excel_file = args[1]

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save(output_excel_file)

if __name__ == '__main__':  
    main(sys.argv)